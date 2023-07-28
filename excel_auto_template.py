import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from tkinter import Tk, filedialog


def write_cell_and_format(file_path, sheet_name, row, column, value, font=None, fill=None):
    """ ｾﾙ操作
    Overview:
        指定したExcelﾌｧｲﾙのｼｰﾄに対してセルの値を書き込む
        セルの書式設定を行う
    Args:
        file_path (str): Excelﾌｧｲﾙﾊﾟｽ
        sheet_name (str): ｼｰﾄ名
        row (int): 書き込むｾﾙ行
        column (int): 書き込むｾﾙ列
        value : ｾﾙに書き込む値
        font (openpyxl.styles.Font, optional): ｾﾙのﾌｫﾝﾄ情報
        fill (openpyxl.styles.PatternFill, optional): ｾﾙの書式情報
    Returns:
        None
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    cell = ws.cell(row=row, column=column)
    cell.value = value

    if font:
        cell.font = font

    if fill:
        cell.fill = fill

    filename = filedialog.asksaveasfilename(
        title="名前を付けて保存",
        filetypes=[("Excel", ".xls"), ("Excel", ".xlsx")],  # ﾌｧｲﾙﾌｨﾙﾀ
        initialdir="./",  # 自分自身のﾃﾞｨﾚｸﾄﾘ
        defaultextension="excel"
    )

    wb.save(filename)
    wb.close()


def create_and_insert_chart(file_path, sheet_name, data_range, chart_title, chart_position):
    """ ｸﾞﾗﾌ作成
    Overview:
        ｸﾞﾗﾌの作成と挿入の自動化
    Args:
        file_path (str): Excelﾌｧｲﾙﾊﾟｽ
        sheet_name (str): ｼｰﾄ名
        data_range (taple: int): ｸﾞﾗﾌ表示するﾃﾞｰﾀ範囲配列(開始列, 開始行, 終了列, 終了行)
        chart_title (str): ｸﾞﾗﾌﾀｲﾄﾙ
        chart_position (str) : ｸﾞﾗﾌ挿入位置(例: "A50")
    Returns:
        None
    """

    # Excelﾌｧｲﾙﾛｰﾄﾞ
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # ﾊﾞｰﾁｬｰﾄｵﾌﾞｼｪｸﾄ作成
    chart = BarChart()
    chart.title = chart_title
    chart.x_axis.title = "X軸ラベル"
    chart.y_axis.title = "Y軸ラベル"

    # ﾃﾞｰﾀ範囲とｶﾃｺﾞﾘ範囲を指定(開始列, 開始行, 終了列, 終了行)
    data = Reference(
        ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
    categories = Reference(
        ws, min_col=data_range[0], min_row=data_range[1]+1, max_row=data_range[3])

    # ﾃﾞｰﾀを追加してｶﾃｺﾞﾘを設定
    chart.add_data(data, titles_from_data=True)  # ﾃﾞｰﾀ範囲の最初の行をｶﾃｺﾞﾘとして使用
    chart.set_categories(categories)

    # ﾁｬｰﾄを指定の位置に追加
    ws.add_chart(chart, chart_position)  # 開始列, 開始行

    filename = filedialog.asksaveasfilename(
        title="名前を付けて保存",
        filetypes=[("Excel", ".xls"), ("Excel", ".xlsx")],  # ﾌｧｲﾙﾌｨﾙﾀ
        initialdir="./",  # 自分自身のﾃﾞｨﾚｸﾄﾘ
        defaultextension="excel"
    )

    # 変更を保存してﾌｧｲﾙを閉じる
    wb.save(filename)
    wb.close()


def read_excel(file_path, sheet_name):
    """ Excel読み込み
    Overview:
        指定したExcelﾌｧｲﾙからﾃﾞｰﾀ読み込み
    Args:
        file_path (str): Excelﾌｧｲﾙﾊﾟｽ
        sheet_name (str): ｼｰﾄ名
    Returns:
        pandas.DataFrame: Excelﾌｧｲﾙから読み込まれた DataFrame ｵﾌﾞｼﾞｪｸﾄ
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    return df


def select_file_dialog():
    """ ﾌｧｲﾙ選択ﾀﾞｲｱﾛｸﾞ
    Overview:
        ﾌｧｲﾙをﾀﾞｲｱﾛｸﾞで選択する
    Args:
        None
    Returns:
        str: 選択したﾌｧｲﾙのﾊﾟｽ文字列
    """

    root = Tk()  # Tkinter ｳｲﾝﾄﾞｳ作成
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("", "")])  # ﾊﾟｽ取得

    root.destroy()  # ｳｲﾝﾄﾞｳ破棄

    return file_path


def select_folder_dialog():
    """ ﾌｫﾙﾀﾞ選択ﾀﾞｲｱﾛｸﾞ
    Overview:
        ﾌｫﾙﾀﾞをﾀﾞｲｱﾛｸﾞで選択する
    Args:
        None
    Returns:
        str: 選択したﾌｫﾙﾀﾞのﾊﾟｽ文字列
    """

    root = Tk()  # Tkinter ｳｲﾝﾄﾞｳ作成
    root.withdraw()

    folder_path = filedialog.askdirectory()  # ﾊﾟｽ取得

    root.destroy()  # ｳｲﾝﾄﾞｳ破棄

    return folder_path
