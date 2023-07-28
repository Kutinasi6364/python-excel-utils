import openpyxl
from tkinter import Tk, filedialog

def select_file_dialog():
    """ ﾌｧｲﾙ選択ﾀﾞｲｱﾛｸﾞ

    Overview:
        処理を行うﾌｧｲﾙを選択(複数選択可)
    Args:
        None
    Returns:
        file_path (str) : 選択したﾌｧｲﾙのﾊﾟｽ
    """

    root = Tk()
    root.withdraw()

    file_paths = filedialog.askopenfilenames(filetypes=[("Excel", ".xlsx"), ("Excel", ".xls")])
    root.destroy()

    return file_paths

def convert_excel_case(file_path, case_option):
    """ 大文字小文字変換
    Overview:
        半角英数を大文字 or 小文字 にすべて変換
    Args:
        file_path (str): 変換を行うﾌｧｲﾙﾊﾟｽ
        case_option (int): 0:大文字 1:小文字 選択
    """

    # Excelﾌｧｲﾙ読み込み
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if case_option == 0:
                        cell.value = cell.value.lower()
                    elif case_option == 1:
                        cell.value = cell.value.upper()

    # 変換後のファイルを保存
    if case_option == 0:
        new_file_path = file_path.replace(".xlsx", "_lowercase.xlsx")
    elif case_option == 1:
        new_file_path = file_path.replace(".xlsx", "_uppercase.xlsx")
    
    wb.save(new_file_path)
    wb.close()

file_paths = select_file_dialog()
case_option = 1 # 0:lower, 1:upper

for file_path in file_paths:
    convert_excel_case(file_path, case_option)