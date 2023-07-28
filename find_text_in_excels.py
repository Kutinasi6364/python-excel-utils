import os
import glob
import openpyxl
from tkinter import Tk, filedialog


def select_folder_dialog():
    """ ﾌｫﾙﾀﾞ選択ﾀﾞｲｱﾛｸﾞ
    Overview:
        ﾌｫﾙﾀﾞをﾀﾞｲｱﾛｸﾞで選択する
    Args:
        None
    Returns:
        str: 選択したﾌｫﾙﾀﾞのﾊﾟｽ文字列
    """

    root = Tk()  # Tkinter ｳｲﾝﾄﾞｳ作成
    root.withdraw()

    folder_path = filedialog.askdirectory()  # ﾊﾟｽ取得

    root.destroy()  # ｳｲﾝﾄﾞｳ破棄

    return folder_path


def find_files_with_string(folder_path, search_string):
    matching_files = []

    # ﾌｫﾙﾀﾞ内のExcelﾌｧｲﾙを検索
    excel_files = glob.glob(os.path.join(folder_path, "*.xls")) + \
        glob.glob(os.path.join(folder_path, "*.xlsx"))

    # 各Excelﾌｧｲﾙのﾌｧｲﾙ名をﾁｪｯｸ
    for file_path in excel_files:
        file_name = os.path.basename(file_path)

        # Excel ﾌｧｲﾙを読み込み
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell_value in row:
                    if isinstance(cell_value, str) and search_string in cell_value:
                        matching_files.append(file_name)
                        break
                else:
                    continue
                break

        # print(file_name)
        wb.close()

    return matching_files


folder_path = select_folder_dialog()
search_string = "test"
result = find_files_with_string(folder_path, search_string)
for name in result:
    print("該当ファイル: " + name)
